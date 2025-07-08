"""
Enhanced Excel term extraction module with NLTK integration
"""

import re
import logging
from pathlib import Path
from typing import List, Set, Dict, Tuple
from openpyxl import load_workbook

# NLTK imports with graceful fallback
try:
    import nltk
    from nltk.tokenize import word_tokenize, sent_tokenize
    from nltk.corpus import stopwords
    from nltk.stem import WordNetLemmatizer
    from nltk.tag import pos_tag
    from nltk.chunk import ne_chunk
    from nltk.tree import Tree
    NLTK_AVAILABLE = True
except ImportError:
    NLTK_AVAILABLE = False

from .config import ProcessingConfig, FinancialTerms


class NLTKDownloadManager:
    """Manage NLTK data downloads with offline support"""
    
    def __init__(self):
        self.required_data = [
            'punkt', 'stopwords', 'wordnet', 'averaged_perceptron_tagger', 
            'maxent_ne_chunker', 'words', 'omw-1.4'
        ]
        self.available_data = set()
        
    def ensure_nltk_data(self) -> bool:
        """Check for existing NLTK data without downloading"""
        if not NLTK_AVAILABLE:
            return False
            
        try:
            # Only check for existing data, don't try to download
            for data_name in self.required_data:
                try:
                    # Try to find existing data first
                    if data_name == 'punkt':
                        nltk.data.find('tokenizers/punkt')
                    elif data_name in ['stopwords', 'wordnet', 'words', 'omw-1.4']:
                        nltk.data.find(f'corpora/{data_name}')
                    elif data_name == 'averaged_perceptron_tagger':
                        nltk.data.find('taggers/averaged_perceptron_tagger')
                    elif data_name == 'maxent_ne_chunker':
                        nltk.data.find('chunkers/maxent_ne_chunker')
                    
                    self.available_data.add(data_name)
                    
                except LookupError:
                    # Data not found, but don't try to download in corporate environment
                    continue
                        
                except Exception:
                    # Data exists but can't be loaded - still count as available
                    continue
            
            # Return True if we have at least the essential components
            essential_data = {'punkt', 'stopwords', 'wordnet'}
            available_essential = len(essential_data.intersection(self.available_data))
            
            if available_essential > 0:
                print(f"NLTK: Found {available_essential}/3 essential packages: {essential_data.intersection(self.available_data)}")
                return True
            else:
                print("NLTK: No essential data found. Use 'python scripts/setup_nltk_manual.py' for setup instructions.")
                return False
            
        except Exception as e:
            print(f"NLTK: Initialization failed: {e}")
            return False


class EnhancedFinancialTermExtractor:
    """NLTK-enhanced financial term extractor with advanced customization"""
    
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.nltk_config = config.nltk_config
        self.financial_terms = FinancialTerms()
        self.logger = logging.getLogger(__name__)
        
        # Initialize NLTK components
        self.nltk_manager = NLTKDownloadManager()
        self.nltk_ready = self.nltk_manager.ensure_nltk_data()
        
        if self.nltk_ready:
            # Initialize NLTK components with domain customization
            self._initialize_nltk_components()
            self.logger.info("NLTK components initialized with financial domain customization")
        else:
            # Fallback to basic processing
            self._initialize_basic_components()
            self.logger.warning("NLTK not available, using basic text processing")
    
    def _initialize_nltk_components(self):
        """Initialize NLTK components with financial domain customization"""
        # Lemmatizer with custom exceptions
        self.lemmatizer = WordNetLemmatizer()
        
        # Enhanced stopwords with financial domain awareness
        base_stopwords = set(stopwords.words('english'))
        
        # Remove financial terms that should be preserved
        self.stop_words = base_stopwords - self.nltk_config.financial_terms_to_preserve
        
        # Add domain-specific stopwords to remove
        self.stop_words.update(self.nltk_config.financial_stopwords_to_remove)
        
        # Compile financial entity patterns for faster matching
        self.financial_entity_patterns = [
            re.compile(pattern, re.IGNORECASE) 
            for pattern in self.nltk_config.financial_entity_patterns
        ]
        
        self.logger.info(f"Configured {len(self.stop_words)} stopwords, {len(self.financial_entity_patterns)} entity patterns")
    
    def _initialize_basic_components(self):
        """Initialize basic components without NLTK"""
        # Fallback to basic stopwords
        self.stop_words = {
            'i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves', 'you', 
            'your', 'yours', 'yourself', 'yourselves', 'he', 'him', 'his', 'himself',
            'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'they', 'them',
            'their', 'theirs', 'themselves', 'what', 'which', 'who', 'whom', 'this',
            'that', 'these', 'those', 'am', 'is', 'are', 'was', 'were', 'be', 'been',
            'being', 'have', 'has', 'had', 'having', 'do', 'does', 'did', 'doing',
            'an', 'the', 'and', 'but', 'if', 'or', 'because', 'as', 'until',
            'while', 'of', 'at', 'by', 'for', 'with', 'about', 'against', 'between',
            'into', 'through', 'during', 'before', 'after', 'above', 'below', 'to',
            'from', 'up', 'down', 'in', 'out', 'on', 'off', 'over', 'under', 'again',
            'further', 'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how',
            'all', 'both', 'each', 'few', 'more', 'most', 'other', 'some', 'such',
            'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than', 'too', 'very'
        }
        self.logger.warning("NLTK not available, using basic text processing")
        
    def extract_headers_from_excel(self, file_path: Path) -> List[dict]:
        """Extract header/label cells from Excel file with enhanced NLTK processing"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            headers = []
            file_stats = {
                'file_path': str(file_path),
                'file_name': file_path.name,
                'total_sheets': len(workbook.sheetnames),
                'sheets': {}
            }
            
            for sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]
                    sheet_headers = self._extract_headers_from_sheet(sheet, sheet_name, file_path)
                    headers.extend(sheet_headers)
                    
                    file_stats['sheets'][sheet_name] = {
                        'max_row': sheet.max_row,
                        'max_column': sheet.max_column,
                        'headers_found': len(sheet_headers)
                    }
                    
                except Exception as e:
                    self.logger.warning(f"Failed to process sheet {sheet_name}: {e}")
                    continue
            
            workbook.close()
            
            # Calculate file totals
            file_stats['total_max_rows'] = max([stats['max_row'] for stats in file_stats['sheets'].values()]) if file_stats['sheets'] else 0
            file_stats['total_max_columns'] = max([stats['max_column'] for stats in file_stats['sheets'].values()]) if file_stats['sheets'] else 0
            file_stats['total_headers_found'] = len(headers)
            
            # Enhanced deduplication with NLTK lemmatization
            unique_headers = self._deduplicate_with_lemmatization(headers)
            
            # Add file statistics to each header
            for header_info in unique_headers:
                header_info['file_stats'] = file_stats
            
            return unique_headers
            
        except Exception as e:
            self.logger.error(f"Failed to extract headers from {file_path}: {e}")
            return []
    
    def _extract_headers_from_sheet(self, sheet, sheet_name: str, file_path: Path) -> List[dict]:
        """Extract headers from sheet with enhanced financial context detection"""
        headers = []
        max_rows = min(200, sheet.max_row)
        
        for row_idx in range(1, max_rows + 1):
            row = sheet[row_idx]
            
            for cell in row:
                if self._is_enhanced_header_cell(cell):
                    cleaned_text = self._enhanced_clean_financial_text(str(cell.value))
                    if cleaned_text and len(cleaned_text) > 2:
                        header_info = {
                            'term': cleaned_text,
                            'original_text': str(cell.value),
                            'file_path': str(file_path),
                            'file_name': file_path.name,
                            'sheet_name': sheet_name,
                            'row': cell.row,
                            'column': cell.column,
                            'column_letter': cell.column_letter,
                            'cell_address': f"{cell.column_letter}{cell.row}"
                        }
                        headers.append(header_info)
        
        return headers
    
    def _is_enhanced_header_cell(self, cell) -> bool:
        """Enhanced header detection with configurable thresholds"""
        if not cell.value or not isinstance(cell.value, str):
            return False
            
        text = str(cell.value).strip()
        
        # Configurable length filtering
        if len(text) < self.config.min_term_length or len(text) > self.config.max_term_length:
            return False
            
        # Configurable numeric content filtering
        number_ratio = len(re.findall(r'\d', text)) / len(text)
        if number_ratio > self.config.numeric_content_threshold:
            return False
        
        # Enhanced financial context detection
        if self.nltk_ready and self.config.require_financial_context:
            financial_score = self._calculate_financial_score_nltk(text)
            if financial_score >= self.config.financial_context_threshold:
                return True
            elif financial_score >= self.nltk_config.financial_relevance_threshold:
                return True
        
        # Fallback to pattern-based detection
        return self._pattern_based_header_detection(text)
    
    def _calculate_financial_score_nltk(self, text: str) -> float:
        """Enhanced financial relevance scoring with domain customization"""
        try:
            # Tokenize and POS tag
            tokens = word_tokenize(text.lower())
            pos_tags = pos_tag(tokens) if self.nltk_config.use_pos_tagging else [(t, 'NN') for t in tokens]
            
            score = 0.0
            total_tokens = len(tokens)
            
            if total_tokens == 0:
                return 0.0
            
            # Enhanced financial keyword scoring
            strong_matches = sum(1 for token, _ in pos_tags 
                               if token in self.financial_terms.strong_financial_indicators)
            weak_matches = sum(1 for token, _ in pos_tags 
                             if token in self.financial_terms.weak_financial_indicators)
            
            # Weighted scoring for financial indicators
            score += (strong_matches / total_tokens) * 0.8  # Strong indicators get high weight
            score += (weak_matches / total_tokens) * 0.4    # Weak indicators get moderate weight
            
            # POS tag-based scoring with domain priorities
            if self.nltk_config.use_pos_tagging:
                pos_score = 0.0
                for token, pos in pos_tags:
                    if pos in self.nltk_config.important_pos_tags:
                        pos_score += self.nltk_config.important_pos_tags[pos]
                
                score += (pos_score / total_tokens) * 0.3
            
            # Enhanced class identifier detection
            class_bonus = 0.0
            for token, _ in pos_tags:
                if token in ['class', 'tranche', 'series', 'tier']:
                    class_bonus += 0.4
                elif re.match(r'^[a-f]$', token):  # Single letter class identifiers
                    class_bonus += 0.3
            
            score += min(class_bonus, 0.6)  # Cap class bonus
            
            # Financial entity pattern matching
            pattern_matches = sum(1 for pattern in self.financial_entity_patterns 
                                if pattern.search(text))
            if pattern_matches > 0:
                score += min(pattern_matches * 0.2, 0.4)  # Cap pattern bonus
            
            # Named Entity Recognition enhancement
            if self.nltk_config.use_named_entity_recognition:
                try:
                    chunked = ne_chunk(pos_tags)
                    financial_entities = self._extract_financial_entities(chunked)
                    if financial_entities:
                        score += min(len(financial_entities) * 0.15, 0.3)
                except:
                    pass
            
            return min(score, 1.0)
            
        except Exception as e:
            self.logger.debug(f"Enhanced NLTK scoring failed for '{text}': {e}")
            return 0.0
    
    def _extract_financial_entities(self, chunked_tokens) -> List[str]:
        """Extract financial entities from NER-chunked tokens"""
        entities = []
        
        for chunk in chunked_tokens:
            if hasattr(chunk, 'label'):  # It's a named entity
                entity_text = ' '.join([token for token, pos in chunk.leaves()])
                entities.append(entity_text)
        
        return entities
    
    def _enhanced_clean_financial_text(self, text: str) -> str:
        """Enhanced text cleaning with advanced NLTK customization"""
        if not text:
            return ""
        
        # Basic cleaning (same as before)
        text = text.lower().strip()
        text = re.sub(r'\d+\.?\d*\s*%', '', text)
        text = re.sub(r'\$[\d,]+\.?\d*', '', text)
        text = re.sub(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b', '', text)
        text = re.sub(r'\([^)]*\)', '', text)
        text = re.sub(r'^\{\d+\}\s*', '', text)
        text = re.sub(r'[:\.,;]+$', '', text)
        
        if self.nltk_ready:
            return self._advanced_nltk_cleaning(text)
        else:
            return self._basic_cleaning(text)
    
    def _advanced_nltk_cleaning(self, text: str) -> str:
        """Advanced NLTK cleaning with financial domain intelligence"""
        try:
            # Tokenize
            tokens = word_tokenize(text)
            
            # POS tagging for intelligent processing
            pos_tags = pos_tag(tokens) if self.nltk_config.use_pos_tagging else [(t, 'NN') for t in tokens]
            
            # Advanced lemmatization with domain customization
            cleaned_tokens = []
            
            for token, pos in pos_tags:
                # Skip very short words except critical financial identifiers
                if len(token) < self.config.min_term_length and token not in self.nltk_config.financial_terms_to_preserve:
                    continue
                
                # Skip stopwords except preserved financial terms
                if token in self.stop_words and token not in self.nltk_config.financial_terms_to_preserve:
                    continue
                
                # Apply custom lemmatization
                if self.nltk_config.use_lemmatization:
                    lemmatized_token = self._custom_lemmatize(token, pos)
                    cleaned_tokens.append(lemmatized_token)
                else:
                    cleaned_tokens.append(token)
            
            # Remove consecutive duplicates while preserving order
            deduplicated = []
            prev_token = None
            for token in cleaned_tokens:
                if token != prev_token:
                    deduplicated.append(token)
                prev_token = token
            
            return ' '.join(deduplicated)
            
        except Exception as e:
            self.logger.debug(f"Advanced NLTK cleaning failed: {e}")
            return self._basic_cleaning(text)
    
    def _custom_lemmatize(self, token: str, pos: str) -> str:
        """Custom lemmatization with financial domain exceptions"""
        # Check for custom exceptions first
        if token in self.nltk_config.custom_lemma_exceptions:
            return self.nltk_config.custom_lemma_exceptions[token]
        
        # Preserve certain financial plurals if configured
        if (self.nltk_config.preserve_financial_plurals and 
            token in self.nltk_config.financial_terms_to_preserve and
            token.endswith('s')):
            return token  # Keep original form
        
        # Standard lemmatization with POS context
        wordnet_pos = self._get_wordnet_pos(pos)
        return self.lemmatizer.lemmatize(token, wordnet_pos)
    
    def _basic_cleaning(self, text: str) -> str:
        """Basic text cleaning without NLTK"""
        # Normalize separators
        text = re.sub(r'[_\-\s]+', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        
        # Simple tokenization
        words = text.split()
        
        # Remove stopwords and filter
        financial_keepers = {'fee', 'tax', 'ytd', 'apr', 'apy', 'cpr', 'psa', 'a', 'b', 'c', 'd', 'e', 'f'}
        cleaned_words = []
        
        for word in words:
            if (len(word) > 2 or word in financial_keepers) and word not in self.stop_words:
                cleaned_words.append(word)
        
        # Remove consecutive duplicates
        deduplicated = []
        prev_word = None
        for word in cleaned_words:
            if word != prev_word:
                deduplicated.append(word)
            prev_word = word
        
        return ' '.join(deduplicated)
    
    def _deduplicate_with_lemmatization(self, headers: List[dict]) -> List[dict]:
        """Smart deduplication using NLTK lemmatization"""
        if not self.nltk_ready:
            # Fallback to simple deduplication
            seen_terms = {}
            unique_headers = []
            for header_info in headers:
                term = header_info['term']
                if term not in seen_terms:
                    seen_terms[term] = True
                    unique_headers.append(header_info)
            return unique_headers
        
        # NLTK-enhanced deduplication
        seen_lemmatized = {}
        unique_headers = []
        
        for header_info in headers:
            term = header_info['term']
            
            # Create lemmatized signature for comparison
            try:
                tokens = word_tokenize(term.lower())
                lemmatized_tokens = [self.lemmatizer.lemmatize(token) for token in tokens]
                lemmatized_signature = ' '.join(sorted(lemmatized_tokens))
                
                if lemmatized_signature not in seen_lemmatized:
                    seen_lemmatized[lemmatized_signature] = True
                    unique_headers.append(header_info)
                    
            except Exception as e:
                # Fallback to exact matching
                if term not in seen_lemmatized:
                    seen_lemmatized[term] = True
                    unique_headers.append(header_info)
        
        return unique_headers


# Maintain backward compatibility
FinancialTermExtractor = EnhancedFinancialTermExtractor