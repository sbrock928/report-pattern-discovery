"""
Excel report generation module for Financial Pattern Discovery System
"""

import logging
from pathlib import Path
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import ProcessingConfig


class ExcelReportGenerator:
    """Generate comprehensive Excel reports"""
    
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)
        
    def generate_report(self, results: Dict[str, Any], output_path: Path):
        """Generate comprehensive Excel report"""
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets
        self._create_summary_sheet(workbook, results)
        self._create_file_statistics_sheet(workbook, results)  # New tab
        self._create_cluster_details_sheet(workbook, results)
        self._create_mappings_sheet(workbook, results)
        self._create_statistics_sheet(workbook, results)
        
        # Save workbook
        workbook.save(output_path)
        self.logger.info(f"Report saved to {output_path}")
    
    def _create_summary_sheet(self, workbook: Workbook, results: Dict[str, Any]):
        """Create summary sheet with key metrics"""
        ws = workbook.create_sheet("Summary")
        
        # Headers
        headers = ['Metric', 'Value', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        metrics = [
            ('Total Files Processed', results.get('total_files', 0), 'Number of Excel files processed'),
            ('Total Terms Extracted', results.get('total_terms', 0), 'Total financial terms found'),
            ('Number of Clusters', results.get('n_clusters', 0), 'Number of term clusters identified'),
            ('Average Cluster Size', results.get('avg_cluster_size', 0), 'Average terms per cluster'),
            ('Silhouette Score', results.get('silhouette_score', 0), 'Clustering quality metric'),
            ('High Confidence Mappings', results.get('high_confidence_count', 0), 'Mappings with high confidence'),
            ('Processing Time', results.get('processing_time', 0), 'Total processing time (seconds)')
        ]
        
        for row, (metric, value, desc) in enumerate(metrics, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=desc)
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_file_statistics_sheet(self, workbook: Workbook, results: Dict[str, Any]):
        """Create file statistics sheet showing rows/columns per input file"""
        ws = workbook.create_sheet("File Statistics")
        
        # File-level summary first
        ws.cell(row=1, column=1, value="File Summary").font = Font(bold=True, size=14)
        
        # File summary headers
        file_headers = ['File Name', 'Total Sheets', 'Max Rows', 'Max Columns', 'Headers Found']
        for col, header in enumerate(file_headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # File summary data
        row = 4
        for file_path, file_stats in results.get('file_statistics', {}).items():
            file_name = Path(file_path).name
            ws.cell(row=row, column=1, value=file_name)
            ws.cell(row=row, column=2, value=file_stats.get('total_sheets', 0))
            ws.cell(row=row, column=3, value=file_stats.get('total_max_rows', 0))
            ws.cell(row=row, column=4, value=file_stats.get('total_max_columns', 0))
            ws.cell(row=row, column=5, value=file_stats.get('total_headers_found', 0))
            row += 1
        
        # Add some spacing
        row += 2
        
        # Sheet-level details
        ws.cell(row=row, column=1, value="Sheet Details").font = Font(bold=True, size=14)
        row += 2
        
        # Sheet detail headers
        sheet_headers = ['File Name', 'Sheet Name', 'Rows', 'Columns', 'Headers Found']
        for col, header in enumerate(sheet_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row += 1
        
        # Sheet detail data
        for file_path, file_stats in results.get('file_statistics', {}).items():
            file_name = Path(file_path).name
            for sheet_name, sheet_stats in file_stats.get('sheets', {}).items():
                ws.cell(row=row, column=1, value=file_name)
                ws.cell(row=row, column=2, value=sheet_name)
                ws.cell(row=row, column=3, value=sheet_stats.get('max_row', 0))
                ws.cell(row=row, column=4, value=sheet_stats.get('max_column', 0))
                ws.cell(row=row, column=5, value=sheet_stats.get('headers_found', 0))
                row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_cluster_details_sheet(self, workbook: Workbook, results: Dict[str, Any]):
        """Create detailed cluster information sheet"""
        ws = workbook.create_sheet("Cluster Details")
        
        # Headers
        headers = ['Cluster ID', 'Canonical Name', 'Term Count', 'Terms', 'Top Features']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows
        row = 2
        for cluster_id, cluster_data in results.get('clusters', {}).items():
            canonical_name = results.get('canonical_names', {}).get(cluster_id, 'unknown')
            
            ws.cell(row=row, column=1, value=cluster_id)
            ws.cell(row=row, column=2, value=canonical_name)
            ws.cell(row=row, column=3, value=cluster_data['count'])
            ws.cell(row=row, column=4, value=', '.join(cluster_data['terms']))
            ws.cell(row=row, column=5, value=', '.join(cluster_data.get('top_features', [])))
            
            row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_mappings_sheet(self, workbook: Workbook, results: Dict[str, Any]):
        """Create term mappings sheet"""
        ws = workbook.create_sheet("Term Mappings")
        
        # Headers - Added location information columns
        headers = [
            'Original Term', 'Canonical Name', 'Cluster ID', 'Fuzzy Score', 'Confidence', 
            'Source File', 'Sheet Name', 'Cell Address', 'Row', 'Column', 'Original Text'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data rows - Sort by fuzzy score descending
        mappings = results.get('mappings', [])
        # Sort mappings by fuzzy_score in descending order (highest confidence first)
        sorted_mappings = sorted(mappings, key=lambda x: x.get('fuzzy_score', 0), reverse=True)
        
        for row, mapping in enumerate(sorted_mappings, 2):
            ws.cell(row=row, column=1, value=mapping['original_term'])
            ws.cell(row=row, column=2, value=mapping['canonical_name'])
            ws.cell(row=row, column=3, value=mapping['cluster_id'])
            ws.cell(row=row, column=4, value=mapping['fuzzy_score'])
            ws.cell(row=row, column=5, value=mapping['confidence'])
            ws.cell(row=row, column=6, value=mapping.get('source_file', 'unknown'))
            ws.cell(row=row, column=7, value=mapping.get('sheet_name', 'unknown'))
            ws.cell(row=row, column=8, value=mapping.get('cell_address', 'unknown'))
            ws.cell(row=row, column=9, value=mapping.get('row', 'unknown'))
            ws.cell(row=row, column=10, value=mapping.get('column', 'unknown'))
            ws.cell(row=row, column=11, value=mapping.get('original_text', 'unknown'))

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_statistics_sheet(self, workbook: Workbook, results: Dict[str, Any]):
        """Create statistics and frequency analysis sheet"""
        ws = workbook.create_sheet("Statistics")
        
        # Term frequency analysis
        ws.cell(row=1, column=1, value="Term Frequency Analysis").font = Font(bold=True, size=14)
        
        headers = ['Term', 'Frequency', 'Percentage', 'Cluster ID']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Calculate term frequencies
        term_freq = {}
        total_terms = 0
        
        for cluster_id, cluster_data in results.get('clusters', {}).items():
            for term in cluster_data['terms']:
                term_freq[term] = term_freq.get(term, 0) + 1
                total_terms += 1
        
        # Sort by frequency
        sorted_terms = sorted(term_freq.items(), key=lambda x: x[1], reverse=True)
        
        # Add data rows
        for row, (term, freq) in enumerate(sorted_terms, 4):
            percentage = (freq / total_terms * 100) if total_terms > 0 else 0
            
            # Find cluster ID for this term
            cluster_id = None
            for cid, cluster_data in results.get('clusters', {}).items():
                if term in cluster_data['terms']:
                    cluster_id = cid
                    break
            
            ws.cell(row=row, column=1, value=term)
            ws.cell(row=row, column=2, value=freq)
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            ws.cell(row=row, column=4, value=cluster_id)
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width